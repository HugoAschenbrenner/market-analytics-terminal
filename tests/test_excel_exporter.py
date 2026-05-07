from io import BytesIO
from datetime import date

from openpyxl import load_workbook

from engines.fixed_income_engine import (
    calculate_bond_risk_metrics,
    calculate_dv01_by_bucket,
    calculate_scenario_pnl,
    generate_fixed_income_commentary,
    load_bond_data,
    portfolio_summary_to_dict,
    summarize_portfolio,
)
from reports.excel_exporter import generate_fixed_income_risk_report


def test_fixed_income_excel_report_generates_valid_workbook():
    bonds = load_bond_data("data/sample_bonds.csv")
    risk_df = calculate_bond_risk_metrics(bonds, valuation_date=date(2026, 5, 6))
    summary = summarize_portfolio(risk_df)
    summary_dict = portfolio_summary_to_dict(summary)
    bucket_df = calculate_dv01_by_bucket(risk_df)
    scenario_df = calculate_scenario_pnl(risk_df)
    commentary = generate_fixed_income_commentary(risk_df, bucket_df, scenario_df)

    report_bytes = generate_fixed_income_risk_report(
        summary=summary_dict,
        risk_df=risk_df,
        bucket_df=bucket_df,
        scenario_df=scenario_df,
        commentary=commentary,
    )

    assert isinstance(report_bytes, bytes)
    assert len(report_bytes) > 0

    workbook = load_workbook(BytesIO(report_bytes), read_only=True)

    expected_sheets = {
        "Summary",
        "Bond_Level_Risk",
        "DV01_Buckets",
        "Scenario_PnL",
        "Methodology",
    }

    assert expected_sheets.issubset(set(workbook.sheetnames))
