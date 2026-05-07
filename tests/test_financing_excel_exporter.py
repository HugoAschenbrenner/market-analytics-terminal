from io import BytesIO
from datetime import date

from openpyxl import load_workbook

from engines.repo_engine import (
    calculate_margin_call,
    calculate_margin_stress_table,
    calculate_repo_sensitivity_table,
    calculate_repo_trade,
    generate_repo_margin_commentary,
    margin_result_to_dict,
    repo_result_to_dict,
)
from engines.sec_lending_engine import (
    calculate_borrow_fee_comparison_table,
    calculate_securities_lending_trade,
    generate_sec_lending_commentary,
    sec_lending_result_to_dict,
)
from reports.excel_exporter import generate_financing_margin_report


def test_financing_margin_excel_report_generates_valid_workbook():
    repo_result = calculate_repo_trade(
        collateral_market_value=10_000_000,
        haircut=0.02,
        repo_rate=0.04,
        start_date=date(2026, 5, 6),
        end_date=date(2026, 6, 5),
        day_count_basis=360,
        currency="EUR",
    )

    repo_sensitivity_df = calculate_repo_sensitivity_table(
        collateral_market_value=10_000_000,
        haircut=0.02,
        repo_rate=0.04,
        start_date=date(2026, 5, 6),
        end_date=date(2026, 6, 5),
        day_count_basis=360,
        currency="EUR",
    )

    margin_result = calculate_margin_call(
        collateral_market_value=repo_result.collateral_market_value,
        cash_amount=repo_result.cash_amount,
        original_haircut=repo_result.haircut,
        collateral_price_shock=-0.05,
        new_haircut=0.04,
    )

    margin_stress_df = calculate_margin_stress_table(
        collateral_market_value=repo_result.collateral_market_value,
        cash_amount=repo_result.cash_amount,
        original_haircut=repo_result.haircut,
    )

    repo_commentary = generate_repo_margin_commentary(margin_result)

    sec_result = calculate_securities_lending_trade(
        security_market_value=5_000_000,
        borrow_fee_rate=0.0125,
        rebate_rate=0.005,
        collateralization_rate=1.02,
        loan_days=30,
        day_count_basis=360,
        utilization_proxy=0.65,
        is_special=False,
    )

    borrow_comparison_df = calculate_borrow_fee_comparison_table(
        security_market_value=5_000_000,
        rebate_rate=0.005,
        collateralization_rate=1.02,
        loan_days=30,
        day_count_basis=360,
    )

    sec_commentary = generate_sec_lending_commentary(sec_result)

    report_bytes = generate_financing_margin_report(
        repo_summary=repo_result_to_dict(repo_result),
        repo_sensitivity_df=repo_sensitivity_df,
        margin_summary=margin_result_to_dict(margin_result),
        margin_stress_df=margin_stress_df,
        repo_commentary=repo_commentary,
        sec_lending_summary=sec_lending_result_to_dict(sec_result),
        borrow_comparison_df=borrow_comparison_df,
        sec_lending_commentary=sec_commentary,
    )

    assert isinstance(report_bytes, bytes)
    assert len(report_bytes) > 0

    workbook = load_workbook(BytesIO(report_bytes), read_only=True)

    expected_sheets = {
        "Repo_Summary",
        "Repo_Sensitivity",
        "Margin_Summary",
        "Margin_Stress",
        "Sec_Lending_Summary",
        "Borrow_Comparison",
        "Methodology",
    }

    assert expected_sheets.issubset(set(workbook.sheetnames))
