from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from engines.portfolio_risk_engine import (
    build_sample_price_data,
    calculate_asset_returns,
    calculate_correlation_matrix,
    calculate_drawdown_series,
    calculate_portfolio_returns,
    calculate_risk_contribution,
    calculate_stress_scenario_table,
    portfolio_risk_summary_to_dict,
    summarize_portfolio_risk,
)
from reports.excel_exporter import generate_portfolio_risk_report


def test_portfolio_risk_excel_report_generates_valid_workbook():
    price_df = build_sample_price_data(n_days=252, seed=42)
    returns_df = calculate_asset_returns(price_df)

    weights = {
        "US_EQUITY": 0.35,
        "EUROPE_EQUITY": 0.20,
        "US_TREASURY": 0.20,
        "IG_CREDIT": 0.15,
        "GOLD": 0.10,
    }

    portfolio_returns = calculate_portfolio_returns(returns_df, weights)
    drawdown = calculate_drawdown_series(portfolio_returns)

    summary = summarize_portfolio_risk(
        returns_df=returns_df,
        weights=weights,
        risk_free_rate=0.02,
    )

    risk_contribution_df = calculate_risk_contribution(returns_df, weights)
    stress_df = calculate_stress_scenario_table(weights)
    correlation_matrix = calculate_correlation_matrix(returns_df)

    weights_df = pd.DataFrame(
        [{"asset": asset, "weight": weight} for asset, weight in weights.items()]
    )

    portfolio_returns_df = portfolio_returns.reset_index()
    portfolio_returns_df.columns = ["date", "portfolio_return"]

    drawdown_df = drawdown.reset_index()
    drawdown_df.columns = ["date", "drawdown"]

    report_bytes = generate_portfolio_risk_report(
        summary=portfolio_risk_summary_to_dict(summary),
        weights_df=weights_df,
        risk_contribution_df=risk_contribution_df,
        correlation_matrix=correlation_matrix,
        stress_df=stress_df,
        portfolio_returns_df=portfolio_returns_df,
        drawdown_df=drawdown_df,
    )

    assert isinstance(report_bytes, bytes)
    assert len(report_bytes) > 0

    workbook = load_workbook(BytesIO(report_bytes), read_only=True)

    expected_sheets = {
        "Portfolio_Summary",
        "Weights",
        "Risk_Contribution",
        "Correlation_Matrix",
        "Stress_Scenarios",
        "Portfolio_Returns",
        "Drawdown",
        "Methodology",
    }

    assert expected_sheets.issubset(set(workbook.sheetnames))
