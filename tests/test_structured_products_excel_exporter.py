from io import BytesIO

from openpyxl import load_workbook

from engines.structured_products_engine import (
    AutocallableTerms,
    build_standard_basket_scenario_paths,
    build_standard_scenario_paths,
    calculate_autocallable_payoff,
    calculate_basket_scenario_table,
    calculate_monte_carlo_results_table,
    calculate_scenario_table,
    payoff_result_to_dict,
    simulate_single_underlying_paths,
    summarize_monte_carlo_results,
)
from reports.excel_exporter import generate_structured_products_report


def test_structured_products_excel_report_generates_valid_workbook():
    terms = AutocallableTerms(
        product_type="Phoenix",
        nominal=1_000_000,
        coupon_rate_per_period=0.02,
        autocall_barrier=0.00,
        coupon_barrier=-0.30,
        protection_barrier=-0.40,
        memory_coupon=True,
    )

    performance_path = [-0.40, 0.02, -0.20, -0.20]
    payoff_result = calculate_autocallable_payoff(terms, performance_path)

    scenario_df = calculate_scenario_table(
        terms=terms,
        scenario_paths=build_standard_scenario_paths(4),
    )

    basket_df = calculate_basket_scenario_table(
        terms=terms,
        basket_scenario_paths=build_standard_basket_scenario_paths(4),
    )

    simulated_paths = simulate_single_underlying_paths(
        number_of_observations=4,
        n_simulations=250,
        volatility=0.25,
        seed=42,
    )

    mc_results_df = calculate_monte_carlo_results_table(
        terms=terms,
        performance_paths=simulated_paths,
    )

    mc_summary = summarize_monte_carlo_results(mc_results_df)

    report_bytes = generate_structured_products_report(
        terms_summary={
            "product_type": terms.product_type,
            "nominal": terms.nominal,
            "coupon_rate_per_period": terms.coupon_rate_per_period,
            "autocall_barrier": terms.autocall_barrier,
            "coupon_barrier": terms.coupon_barrier,
            "protection_barrier": terms.protection_barrier,
            "memory_coupon": terms.memory_coupon,
            "performance_path": performance_path,
        },
        custom_path_result=payoff_result_to_dict(payoff_result),
        single_scenario_df=scenario_df,
        basket_scenario_df=basket_df,
        monte_carlo_summary=mc_summary,
        monte_carlo_results_df=mc_results_df,
    )

    assert isinstance(report_bytes, bytes)
    assert len(report_bytes) > 0

    workbook = load_workbook(BytesIO(report_bytes), read_only=True)

    expected_sheets = {
        "Terms",
        "Custom_Path_Result",
        "Single_Scenarios",
        "Worst_Of_Basket",
        "Monte_Carlo_Summary",
        "MC_Distribution",
        "Methodology",
    }

    assert expected_sheets.issubset(set(workbook.sheetnames))
